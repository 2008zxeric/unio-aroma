import openpyxl, json

f_old = '/Users/ericmac/Desktop/\u65e7.xlsx'
f_new = '/Users/ericmac/Desktop/1.xlsx'
f_out = '/Users/ericmac/Desktop/1.xlsx'

# Load 旧.xlsx to extract the product name order (from row 24 onwards, after header at row 23)
wb_old = openpyxl.load_workbook(f_old, data_only=True)
ws_old = wb_old.active
old_rows = ws_old.iter_rows(min_row=24, values_only=True)  # start from y1 data rows

# Build ordered list of (product_name, english_name) pairs from 旧.xlsx
old_order = []
for row in old_rows:
    if row[2] is not None:  # column C = 产品名称
        old_order.append(row[2])

print(f"旧.xlsx 产品数量: {len(old_order)}")
print(f"前5个: {old_order[:5]}")

# Load 1.xlsx
wb_new = openpyxl.load_workbook(f_new)
ws_new = wb_new.active

# Get all data rows (row 1 is header)
header = [cell for cell in ws_new[1]]
data_rows = []
for row in ws_new.iter_rows(min_row=2, values_only=False):
    if any(cell.value is not None for cell in row):
        data_rows.append(row)

print(f"\n1.xlsx 数据行数: {len(data_rows)}")

# Build a map: product_name -> row object for exact matching
name_to_row = {}
for r in data_rows:
    name = r[1].value  # column B = 产品名称
    if name:
        name_to_row[name] = r

# Sort based on old_order, appending extras at the end
sorted_rows = []
extra_rows = []
for name in old_order:
    if name in name_to_row:
        sorted_rows.append(name_to_row[name])
        del name_to_row[name]
    # if not in 1.xlsx, skip (not present in new file)

# Append remaining (not in old order, but present in 1.xlsx)
for name, r in name_to_row.items():
    extra_rows.append(r)

all_sorted = sorted_rows + extra_rows

print(f"匹配到旧顺序的行: {len(sorted_rows)}")
print(f"额外行(放最后): {len(extra_rows)}")
if extra_rows:
    print(f"额外产品: {[r[1].value for r in extra_rows]}")

# Write back to workbook
# Clear existing rows (keep header at row 1)
for row in ws_new.iter_rows(min_row=2):
    for cell in row:
        cell.value = None

# Write sorted data
for i, row_obj in enumerate(all_sorted, start=2):
    values = [cell.value for cell in row_obj]
    for j, val in enumerate(values):
        ws_new.cell(row=i, column=j+1, value=val)

wb_new.save(f_out)
print(f"\n已保存到: {f_out}")
