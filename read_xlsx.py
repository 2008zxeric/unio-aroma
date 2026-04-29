import openpyxl, json

f1 = '/Users/ericmac/Desktop/\u65e7.xlsx'
f2 = '/Users/ericmac/Desktop/1.xlsx'

result = {}
for label, path in [("\u65e7", f1), ("1", f2)]:
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        result[label] = {
            "sheet": ws.title,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "rows_1_3": [list(r) for r in ws.iter_rows(min_row=1, max_row=3, values_only=True)]
        }
    except Exception as e:
        result[label] = {"error": str(e)}

print(json.dumps(result, ensure_ascii=False, default=str))
