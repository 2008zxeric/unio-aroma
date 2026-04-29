import openpyxl, json

f1 = '/Users/ericmac/Desktop/\u65e7.xlsx'
f2 = '/Users/ericmac/Desktop/1.xlsx'

def read_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        rows.append(list(row))
    return {"sheet": ws.title, "max_row": ws.max_row, "max_col": ws.max_column, "rows": rows}

result = {}
for label, path in [("\u65e7", f1), ("1", f2)]:
    try:
        result[label] = read_xlsx(path)
    except Exception as e:
        result[label] = {"error": str(e)}

with open("/Users/EricMac/.qclaw/workspace/xlsx_data.json", "w", encoding="utf-8") as f:
    json.dump(result, f, ensure_ascii=False, default=str)

print("Done, written to xlsx_data.json")
print("\u65e7 rows:", result.get("\u65e7", {}).get("max_row", "?"))
print("1 rows:", result.get("1", {}).get("max_row", "?"))
